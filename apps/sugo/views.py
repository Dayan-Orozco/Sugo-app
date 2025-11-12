from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import StreamerForm
from .models import *
from django.db.models import Q, Sum
from django.core.paginator import Paginator
import openpyxl
import csv
from io import TextIOWrapper
from django.contrib import messages
from django.utils.dateparse import parse_datetime, parse_date
import pytz
from datetime import datetime, timedelta
from django.utils.timezone import now
from django.utils.http import urlencode


def home(request):
    return render(request, 'index.html')


###################################################### STREAMERS #########################################################################

def register_streamer(request): # Registro de Streamers
    if request.method == 'POST':
        form = StreamerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('streamer_list')  # redirige a otra vista, puede ser un listado
    else:
        form = StreamerForm()
    return render(request, 'streamer-register.html', {'form': form})

def list_streamer(request): # listar Streamers
    if request.user.groups.exists() and request.user.groups.all()[0].name == "Streamer":
        return redirect('home') 
    
    query = request.GET.get('q', '')
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    order_by = request.GET.get('order_by', 'join_date')  # campo por defecto
    direction = request.GET.get('direction', 'desc')  # 'asc' o 'desc'
    # ordenar
    if direction == 'asc':
        order = order_by
    else:
        order = f'-{order_by}'

    streamers = Streamer.objects.filter(enable=True).order_by(order)
    
    if query:
        streamers = streamers.filter(
            Q(name__icontains=query) |
            Q(real_name__icontains=query) |
            Q(id__icontains=query) | 
            Q(country__icontains=query) |
            Q(status__icontains=query)
        )

    if start_date_param:
        # Si no viene fecha final, poner la de hoy
        if not end_date_param:
            end_date_param = now().date().isoformat()  # pasamos a string

        streamers = streamers.filter(
            join_date__date__gte=parse_date(start_date_param),
            join_date__date__lte=parse_date(end_date_param)
        )

    columns = [
        ('id', 'ID'),
        ('name', 'Apodo'),
        ('real_name', 'Nombre Real'),
        ('join_date', 'Fecha de Registro'),
        ('join_date', 'Tiempo/Agencia'),
        ('status', 'Estado'),
        ('telegram_number', 'Telegram'),
        ('country', 'Pais'),
        ('charm_level', 'Encanto'),
        ('vip_level', 'VIP'),
        ('is_capacite', '¿Capacitada?'),
        ('reuse_acount', 'Reuzada'),
    ]

    paginator = Paginator(streamers, 50)
    pending = streamers.filter(join_date__isnull=True)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    total = int(len(streamers)) - int(len(pending))
    return render(request, 'streamer-list.html', {
        'page_obj': page_obj,
        'query': query,
        'order_by': order_by,
        'direction': direction,
        'total': total,
        'pending': pending.count,
        'columns': columns,
    })

def edit_streamer(request, pk): # Editar Streamers
    streamer = get_object_or_404(Streamer, pk=pk)
    if request.method == 'POST':
        form = StreamerForm(request.POST, instance=streamer)
        if form.is_valid():
            form.save()
            return redirect('streamer_list')
    else:
        form = StreamerForm(instance=streamer)
    return render(request, 'streamer-edit.html', {'form': form})

def delete_streamer(request, pk):   # Eliminar Streamers
    streamer = get_object_or_404(Streamer, pk=pk)
    streamer.enable = False
    streamer.save()
    return redirect('streamer_list')

# VISTA PARA SUBIR EXCEL DE CHICAS
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        print("Se hizo post upload_excel")
        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser formato .xlsx')
            return redirect('upload_excel')

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # print("Chica", row[0], row[1]," - ", row[2], row[3])
            # print(row[7], row[8])
            streamer_id = row[0]
            name = row[1]
            real_name = row[2]
            join_date = row[3]
            telegram_number = str(row[4])
            whatsapp_number = str(row[5])
            country = row[6]
            charm_level = int(row[7])
            vip_level = int(row[8])
            # print(name)
            name = str(name).strip() if name else ''
            real_name = str(real_name).strip() if real_name else ''

            if not real_name:
                real_name = name

            # print("real_name", real_name,"name", name)
            # Convertir join_date si es texto
            if isinstance(join_date, str):
                try:
                    join_date = datetime.strptime(join_date, '%Y-%m-%d').date()
                except ValueError:
                    join_date = datetime.now().date()

            # print("join_date", join_date, type(join_date))
            try:
                streamer = Streamer.objects.get(id=streamer_id)
                # print("Actualizando")
                created = False
            except Streamer.DoesNotExist:
                streamer = Streamer(id=streamer_id)
                # print("Creando")
                created = True

            # Asignar o actualizar campos
            streamer.real_name = real_name
            streamer.name = name
            streamer.join_date = join_date
            streamer.telegram_number = telegram_number
            streamer.whatsapp_number = whatsapp_number
            streamer.country = country
            streamer.charm_level = charm_level
            streamer.vip_level = vip_level

            # Guardar para activar signals
            streamer.save()    
        messages.success(request, 'Chicas registradas correctamente desde el Excel.')
    return render(request, 'streamer-upload-excel.html')

# VISTA QUE AGREGA FECHA DE QUE SE CAPACITO UNA STREAMERS 
def capacitar_streamer(request, pk):
    streamer = get_object_or_404(Streamer, pk=pk)
    streamer.is_capacite = True
    streamer.date_capacite = timezone.now()
    streamer.save()
    return redirect('streamer_list') 

def set_join_date(request, pk): # CONFIRMA LA FECHA DE INGRESO
    if request.method == "POST":
        fecha_str = request.POST.get('fecha')
        streamer = get_object_or_404(Streamer, id=pk)

        if fecha_str:
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d %H:%M:%S")
                fecha = timezone.make_aware(fecha, timezone.get_current_timezone())
                streamer.join_date = fecha
                streamer.save()
                print("Se agrego")
            except ValueError as e:
                print("Error", e)
    return redirect('streamer_list')  # Cambia por el nombre real de tu vista/listado

###################################################### DAILY PERFORMANS #########################################################################
def list_daily_performance(request):
    if request.user.groups.exists() and request.user.groups.all()[0].name == "Streamer":
        return redirect('home')

    # Parámetros
    query = request.GET.get('q', '')
    start_date_param = request.GET.get('start_date', '')
    end_date_param = request.GET.get('end_date', '')
    quick_filter = request.GET.get('range', '')  # 7, 15 o 30
    order_by = request.GET.get('order_by', 'date')
    direction = request.GET.get('direction', 'desc')
    view_mode = request.GET.get('view', 'individual')  # individual o total

    # Ordenamiento dinámico
    order = order_by if direction == 'asc' else f'-{order_by}'

    performances = DailyPerformance.objects.all()

    # FILTRO PRINCIPAL (mutuamente exclusivo)
    if start_date_param:  
        # Prioridad a fechas manuales
        if not end_date_param:
            end_date_param = now().date().isoformat()
        performances = performances.filter(
            date__gte=parse_date(start_date_param),
            date__lte=parse_date(end_date_param)
        )

    elif quick_filter in ['7', '15', '30']:
        # Si no hay fechas manuales, aplicar filtro rápido
        days = int(quick_filter)+1
        start_date = now().date() - timedelta(days=days)
        performances = performances.filter(date__gte=start_date)

    else:
        # Si no hay nada, últimos 30 días por defecto
        performances = performances.filter(date__gte=now().date() - timedelta(days=31))

    # Filtro por búsqueda
    if query:
        performances = performances.filter(
            Q(streamer_name__icontains=query) |
            Q(streamer__id__icontains=query)
        )

    # AGRUPAR SI ES "TOTAL"
    if view_mode == "total":
        performances = (
            performances
            .values('streamer__id', 'streamer_name')
            .annotate(
                total_diamonds=Sum('total_diamonds'),
                diamonds_chat=Sum('diamonds_chat'),
                diamonds_gifts=Sum('diamonds_gifts'),
                diamonds_video=Sum('diamonds_video')
            )
            .order_by(order.replace('date', 'streamer__id'))  # date no existe en este modo
        )
    else:
        performances = performances.order_by(order)

    # Totales
    totals = performances.aggregate(
        total_diamonds_sum=Sum('total_diamonds'),
        diamonds_chat_sum=Sum('diamonds_chat'),
        diamonds_gifts_sum=Sum('diamonds_gifts'),
        diamonds_video_sum=Sum('diamonds_video')
    )

    # Conversión a USD
    totals['total_diamonds_usd'] = (totals['total_diamonds_sum'] or 0) / 5000
    totals['diamonds_chat_usd'] = (totals['diamonds_chat_sum'] or 0) / 5000
    totals['diamonds_gifts_usd'] = (totals['diamonds_gifts_sum'] or 0) / 5000
    totals['diamonds_video_usd'] = (totals['diamonds_video_sum'] or 0) / 5000

    columns = [
        ('date', 'Fecha'),
        ('streamer', 'ID'),
        ('streamer_name', 'Streamer'),
        ('total_diamonds', 'Diamantes Totales'),
        ('diamonds_chat', 'Diamantes Chat'),
        ('diamonds_gifts', 'Diamantes Regalos'),
        ('diamonds_video', 'Diamantes Video'),
    ]    
    view_modes = [
        ('individual', 'Individual'),
        ('total', 'Total')
    ]

    paginator = Paginator(performances, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'daily-list.html', {
        'page_obj': page_obj,
        'query': query,
        'order_by': order_by,
        'direction': direction,
        'columns': columns,
        'totals': totals,
        'total': performances.count(),
        'quick_filter': quick_filter,
        'start_date': start_date_param,
        'end_date': end_date_param,       
        'quick_ranges': ['7', '15', '30'],
        'view_mode': view_mode,
        'view_modes': view_modes,  # <- aquí lo mandamos al template
    })


# Vista para subir Excel con rendimiento diario
def upload_excel_daily_performance(request):
    if request.method == 'POST' and request.FILES.get('data_file'):
        data_file = request.FILES['data_file']

        file_name = data_file.name.lower()
        is_excel = file_name.endswith('.xlsx')
        is_csv = file_name.endswith('.csv')

        if not (is_excel or is_csv):
            messages.error(request, 'El archivo debe ser formato .xlsx o .csv')
            return redirect('upload_excel_daily_performance')

        rows = []
        if is_excel:
            wb = openpyxl.load_workbook(data_file)
            sheet = wb.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
        elif is_csv:
            # Abrimos el CSV en modo texto
            wrapper = TextIOWrapper(data_file.file, encoding='utf-8')
            reader = csv.reader(wrapper)
            next(reader, None)  # Saltar encabezado
            rows = list(reader)

        # Se espera un encabezado en la fila 1
        for row in rows:
            # Columnas esperadas:
            # [date, streamer_name, streamer_id, total_diamonds, diamonds_chat, diamonds_gifts, diamonds_video]
            date_value = row[0]
            streamer_name = row[1]
            streamer_id = row[2]
            total_diamonds = row[7]
            diamonds_chat = row[8]
            diamonds_gifts = row[9]
            diamonds_video = row[10]

            # Limpieza de datos
            streamer_name = str(streamer_name).strip() if streamer_name else ''

            # Convertir fecha
            if isinstance(date_value, str):
                try:
                    date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
                except ValueError:
                    date_value = datetime.now().date()
            elif isinstance(date_value, datetime):
                date_value = date_value.date()

            # Buscar streamer
            try:
                streamer = Streamer.objects.get(id=streamer_id)
                # Si el nombre cambió, actualizarlo
                if streamer_name and streamer.name != streamer_name:
                    streamer.name = streamer_name
                    streamer.save(update_fields=['name'])

            except Streamer.DoesNotExist:
                print("Este usuario no existe", streamer_id)
                data = Streamer(id=streamer_id, name=streamer_name, real_name=streamer_name, vip_level=0, status="", charm_level=0, enable=False)
                data.save()
                streamer = data
                pass

            # Crear o actualizar registro
            DailyPerformance.objects.update_or_create(
                streamer=streamer,
                date=date_value,
                defaults={
                    'streamer_name': streamer_name,
                    'total_diamonds': total_diamonds or 0,
                    'diamonds_chat': diamonds_chat or 0,
                    'diamonds_gifts': diamonds_gifts or 0,
                    'diamonds_video': diamonds_video or 0,
                }
            )
            
            print("Creado")

        messages.success(request, 'Rendimientos diarios cargados correctamente desde el Excel.')

    return render(request, 'daily-upload-excel.html')
